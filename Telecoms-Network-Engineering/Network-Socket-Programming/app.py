from flask import Flask, request, render_template
import os
import re
from PyPDF2 import PdfReader
from bs4 import BeautifulSoup
import openpyxl

app = Flask(__name__)

# Directories for different file types
TEXT_FILES_DIR = "text_files"
PDF_FILES_DIR = "pdf_files"
HTML_FILES_DIR = "html_files"
EXCEL_FILES_DIR = "excel_files"

@app.route('/')
def home():
    """
    Home route with a form for entering a keyword.
    """
    return render_template("index.html")

@app.route('/search', methods=['POST'])
def search():
    """
    Search for a keyword in the selected file types.
    """
    keyword_query = request.form.get('keyword', '').strip()
    if not keyword_query:
        return "Please enter a valid keyword.", 400

    # Check if regex search is enabled
    use_regex = 'use_regex' in request.form

    # Get selected file types
    file_types = request.form.getlist('file_types')
    if not file_types:
        return "Please select at least one file type to search.", 400

    results = []

    def match_text(text):
        """
        Check if the text matches the keyword or regex.
        """
        if use_regex:
            try:
                return re.search(keyword_query, text, re.IGNORECASE) is not None
            except re.error:
                return False  # Return False if regex is invalid
        else:
            return keyword_query.lower() in text.lower()

    # Search in text files
    if 'txt' in file_types:
        for filename in os.listdir(TEXT_FILES_DIR):
            file_path = os.path.join(TEXT_FILES_DIR, filename)
            if os.path.isfile(file_path) and filename.endswith('.txt'):
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        lines = file.readlines()
                        for line_number, line in enumerate(lines, start=1):
                            if match_text(line):
                                results.append({
                                    "file": filename,
                                    "line_number": line_number,
                                    "context": line.strip()
                                })
                except Exception as e:
                    return f"Error opening text file {filename}: {e}", 500

    # Search in PDF files
    if 'pdf' in file_types:
        for filename in os.listdir(PDF_FILES_DIR):
            file_path = os.path.join(PDF_FILES_DIR, filename)
            if os.path.isfile(file_path) and filename.endswith('.pdf'):
                try:
                    with open(file_path, 'rb') as file:
                        pdf_reader = PdfReader(file)
                        for page_num, page in enumerate(pdf_reader.pages):
                            text = page.extract_text()
                            if text and match_text(text):
                                results.append({
                                    "file": filename,
                                    "page_number": page_num + 1,
                                    "context": text.strip()[:100]
                                })
                except Exception as e:
                    return f"Error opening PDF file {filename}: {e}", 500

    # Search in HTML files
    if 'html' in file_types:
        for filename in os.listdir(HTML_FILES_DIR):
            file_path = os.path.join(HTML_FILES_DIR, filename)
            if os.path.isfile(file_path) and filename.endswith('.html'):
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        soup = BeautifulSoup(file, 'html.parser')
                        text = soup.get_text()
                        for line_number, line in enumerate(text.splitlines(), start=1):
                            if match_text(line):
                                results.append({
                                    "file": filename,
                                    "line_number": line_number,
                                    "context": line.strip()
                                })
                except Exception as e:
                    return f"Error opening HTML file {filename}: {e}", 500

    # Search in Excel files
    if 'excel' in file_types:
        for filename in os.listdir(EXCEL_FILES_DIR):
            file_path = os.path.join(EXCEL_FILES_DIR, filename)
            if os.path.isfile(file_path) and filename.endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(file_path)
                    for sheet in wb.sheetnames:
                        sheet_obj = wb[sheet]
                        for row in sheet_obj.iter_rows():
                            for cell in row:
                                if cell.value and isinstance(cell.value, str) and match_text(cell.value):
                                    results.append({
                                        "file": filename,
                                        "sheet": sheet,
                                        "cell": cell.coordinate,
                                        "context": cell.value.strip()
                                    })
                except Exception as e:
                    return f"Error opening Excel file {filename}: {e}", 500

    # Return results
    if results:
        return render_template("results.html", keyword=keyword_query, results=results)
    else:
        return f"No results found for the keyword '{keyword_query}'.", 200

if __name__ == "__main__":
    # Ensure directories exist
    for directory in [TEXT_FILES_DIR, PDF_FILES_DIR, HTML_FILES_DIR, EXCEL_FILES_DIR]:
        if not os.path.exists(directory):
            os.makedirs(directory)

    app.run(host="127.0.0.1", port=5000, debug=True)

